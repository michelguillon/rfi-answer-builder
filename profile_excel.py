"""
profile_excel.py — RFI Excel schema profiler  [spec Step 2]
============================================================
Inspects a client-supplied RFI Excel file and produces an approved
column-to-role mapping (config_rfi_<slug>.json) that downstream loaders
will consume. Runs in three phases:

  Phase 1 — schema discovery.
    Open the workbook with openpyxl, pick the sheet most likely to hold
    Q&A data (auto-largest), and for every column produce a profile:
    header text, % non-empty, average word count, question-mark rate,
    cardinality, sample values, and a heuristic role guess.

  Phase 2 — LLM mapping recommendation.
    Send the column profile to Mistral (`mistral-small-latest`) and
    ask for a structured role mapping plus client/date inferences.
    Every API call goes through call_with_retry() so a transient 429
    or 5xx doesn't lose the run.

  Phase 3 — strict validation + human approval.
    Validate the LLM's output BEFORE showing it to the human: every
    role must be in the allowed set, exactly one column must be marked
    `question`, exactly one `answer`. Print the proposal in a readable
    form. On approval, write config_rfi_<slug>.json.

CLI:
    python profile_excel.py data/<file>.xlsx
    # ↑ run via `docker compose run --rm pipeline ...` in this repo.

----------------------------------------------------------------------
ARCHITECTURAL DECISION: discover structure, do not assume it.
RFI files arrive in shapes the team does not control — clients set the
format. Column names, positions, sheet names, and even sheet *layout*
vary between files. Hard-coding "the question is always column B" is
how this project would break the day a new RFI lands. So the profiler
discovers the structure on every file and persists the discovered
mapping. The next pipeline step reads that persisted mapping rather
than re-discovering. See docs/LEARNING_NOTES_RFI.md, entry 2.

ARCHITECTURAL DECISION: three layers (heuristic → LLM → human),
not one. Each layer catches a different failure mode:
  - Heuristic is deterministic and cheap; it gives the LLM a strong
    starting signal and a visible baseline a human can compare against.
    Brittle on weird files where headers are missing or misleading.
  - LLM is robust to phrasing variation ("Q." vs "Question" vs "Item");
    it can also infer client/date from filename. But it occasionally
    violates explicit schema constraints (assigns two columns to
    `question`, invents a role like `q_id` not in the allowed set).
  - Human is the last gate. A misconfigured schema corrupts the entire
    corpus downstream and is silent — wrong answers retrieved against
    misattributed chunks, no exception thrown. The approval prompt is
    load-bearing, not ceremonial. See spec Decision 2.

ARCHITECTURAL DECISION: validate the LLM output BEFORE the human sees it.
A tired human eyeballing a proposal table can miss "this column is
labelled `question` AND `answer`". The validator catches that
mechanically and either drops the proposal or surfaces it as a hard
issue, so the human reviews a sanitised candidate or none at all —
not a subtly-broken one.

ARCHITECTURAL DECISION: openpyxl with `data_only=True`.
RFI answer cells sometimes contain formulas (=A1&" - "&B1, vlookups
into lookup tables, etc.). The semantic content is the *resolved*
value, not the formula string. data_only=True returns the cached
result from the last time the file was saved with Excel/Sheets. The
risk: if a file was never opened in a spreadsheet app since the
formulas were entered, the cache is empty and we see None. Acceptable:
that is a file-quality problem we surface in the heuristic
(% non-empty drops) rather than papering over.

ARCHITECTURAL DECISION: one config file per source file, overwrite
on re-profile. Configs are decoupled, so re-profiling one file cannot
corrupt the mapping for another. Overwrite is the user-requested
behaviour — re-running the profiler is how you correct a bad mapping,
and refusing to overwrite would force a manual delete each time.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from pipeline.mistral_helpers import call_with_retry, get_client


# ─── Constants ──────────────────────────────────────────────────────────
MISTRAL_MODEL = "mistral-small-latest"

# How many rows to profile per column. Capped so a 50k-row file does not
# stall the profiler — we only need enough rows to compute stable stats.
MAX_PROFILE_ROWS = 200
# How many sample values to surface to the LLM and human. Three is enough
# to communicate "what does a typical cell look like" without flooding
# the prompt or the screen.
SAMPLE_VALUES = 3

# Known header tokens. The role inference falls back on stats only when
# the header is unrecognised, so the list does not have to be exhaustive
# — just cover the obvious cases.
QUESTION_HEADERS = {
    "question", "questions", "q", "query", "item", "ask",
    "rfi question", "rfi questions", "topic question",
}
ANSWER_HEADERS = {
    "answer", "answers", "a", "response", "responses", "reply", "replies",
    "our response", "our answer", "utiq response",
}
CONTEXT_HEADERS = {
    "context", "background", "notes", "note", "details",
    "comment", "comments", "supporting evidence", "evidence",
}
# Custom metadata header → suggested snake_case role name.
METADATA_HEADERS: dict[str, str] = {
    "category": "category", "topic": "category", "section": "category",
    "area": "category", "domain": "category", "type": "category",
    "tag": "category", "tags": "category",
    "subcategory": "subcategory", "sub-category": "subcategory",
    "ref": "ref", "reference": "ref", "id": "ref",
    "#": "ref", "number": "ref", "no": "ref", "no.": "ref",
    "client": "client_internal_ref", "client ref": "client_internal_ref",
}

# Core single-instance roles + ignore. Anything else from the LLM is
# treated as a metadata field name (subject to snake_case validation).
RESERVED_ROLES = {"question", "answer", "context", "ignore"}
ROLE_NAME_PATTERN = re.compile(r"^[a-z][a-z0-9_]*$")


# ─── Dataclasses ────────────────────────────────────────────────────────
@dataclass
class ColumnProfile:
    """All the per-column statistics the profiler computes in Phase 1.

    `letter` is the Excel column letter (A, B, ..., AA, ...). We carry
    both the integer index (implicit via position in the parent list)
    and the letter because configs reference columns by letter (the
    human-readable address users see in Excel).
    """
    letter: str
    header: str
    non_empty_count: int
    pct_non_empty: float
    avg_word_count: float
    question_mark_rate: float
    cardinality: int
    sample_values: list[str]
    heuristic_role: str
    heuristic_reason: str


@dataclass
class SheetProfile:
    name: str
    n_data_rows: int
    n_cols: int
    columns: list[ColumnProfile]


@dataclass
class MappingProposal:
    """The LLM's recommendation, post-validation.

    column_roles maps Excel column letter → role string. Roles that are
    in RESERVED_ROLES are treated specially in the config (question/
    answer/context become single keys, ignore becomes a list). Anything
    else is a custom metadata field name and gets a top-level entry in
    `columns` plus an entry in `metadata_fields`.
    """
    sheet: str
    column_roles: dict[str, str]
    client: str | None
    date: str | None
    reasoning: str


# ─── Phase 1: schema discovery ──────────────────────────────────────────
def _truncate(text: str, n: int = 80) -> str:
    """Compact a value for display. Long answer cells are 500+ chars; we
    show a slice with an ellipsis."""
    text = text.replace("\n", " ").replace("\r", " ").strip()
    return text if len(text) <= n else text[: n - 1] + "…"


def _cell_to_str(value: Any) -> str:
    """openpyxl returns native types (int, float, datetime, str, None).
    For profiling we coerce everything to a clean string. Empty cells
    become "" so the % non-empty calculation is straightforward."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def profile_sheet(ws, header_row: int = 1) -> SheetProfile:
    """Walk the sheet and compute per-column statistics.

    ARCHITECTURAL DECISION: header_row is a parameter, not a constant.
    Real RFI files frequently have a preamble before the Q&A table:
    a title row, a "confidential" notice, a form-style metadata block
    ("Company:", "Completed By:") that occupies several rows. If we
    blindly treat row 1 as the header, those preamble rows dominate
    the column statistics and the heuristic mis-classifies everything.
    Exposing header_row as a parameter (and as a `--header-row` CLI
    flag) lets the human override when auto-detect can't find the
    Q&A boundary. Default is 1 because clean Q&A tables are the
    common case.

    We trust ws.max_row / ws.max_column but cap the data window at
    MAX_PROFILE_ROWS for runtime safety.
    """
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if header_row < 1 or header_row > max_row:
        raise ValueError(
            f"header_row={header_row} out of range for sheet "
            f"'{ws.title}' (1..{max_row})"
        )
    first_data_row = header_row + 1
    # Cap the rows we read for stats. We still record the true data row
    # count separately so configs can store it for downstream visibility.
    rows_available = max(max_row - header_row, 0)
    data_window = min(rows_available, MAX_PROFILE_ROWS)

    columns: list[ColumnProfile] = []
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        header = _cell_to_str(ws.cell(row=header_row, column=col_idx).value)
        values: list[str] = []
        for row_idx in range(first_data_row, first_data_row + data_window):
            v = _cell_to_str(ws.cell(row=row_idx, column=col_idx).value)
            if v != "":
                values.append(v)
        non_empty = len(values)
        pct_non_empty = (non_empty / data_window) if data_window else 0.0
        avg_words = (sum(len(v.split()) for v in values) / non_empty) if non_empty else 0.0
        q_marks = sum(1 for v in values if v.rstrip().endswith("?"))
        q_mark_rate = (q_marks / non_empty) if non_empty else 0.0
        cardinality = len(set(values))
        samples = [
            _truncate(v) for v in values[:SAMPLE_VALUES]
        ]
        role, reason = _infer_heuristic_role(
            header=header,
            pct_non_empty=pct_non_empty,
            avg_words=avg_words,
            q_mark_rate=q_mark_rate,
            cardinality=cardinality,
            non_empty=non_empty,
        )
        columns.append(ColumnProfile(
            letter=letter,
            header=header,
            non_empty_count=non_empty,
            pct_non_empty=pct_non_empty,
            avg_word_count=avg_words,
            question_mark_rate=q_mark_rate,
            cardinality=cardinality,
            sample_values=samples,
            heuristic_role=role,
            heuristic_reason=reason,
        ))

    return SheetProfile(
        name=ws.title,
        n_data_rows=max(max_row - header_row, 0),
        n_cols=max_col,
        columns=columns,
    )


def _infer_heuristic_role(
    header: str,
    pct_non_empty: float,
    avg_words: float,
    q_mark_rate: float,
    cardinality: int,
    non_empty: int,
) -> tuple[str, str]:
    """Cheap rule-based role guess. Order matters — earlier rules win.

    The reasoning string explains *which* rule fired, so the LLM and the
    human can both see how the heuristic reached its conclusion. That is
    the difference between a useful pre-classifier and a black box.
    """
    h = header.lower().strip()

    # Rule 1: explicit header match. Strongest signal we have.
    if h in QUESTION_HEADERS:
        return "question", f"header '{header}' matches question pattern"
    if h in ANSWER_HEADERS:
        return "answer", f"header '{header}' matches answer pattern"
    if h in CONTEXT_HEADERS:
        return "context", f"header '{header}' matches context pattern"
    if h in METADATA_HEADERS:
        return METADATA_HEADERS[h], f"header '{header}' matches metadata pattern"

    # Rule 2: mostly empty → ignore. A column that's 95% blank can't be
    # carrying load-bearing question or answer content.
    if pct_non_empty < 0.10:
        return "ignore", f"{pct_non_empty:.0%} non-empty (below 10% threshold)"

    # Rule 3: short text + low cardinality → categorical metadata.
    if avg_words <= 3 and non_empty > 0 and cardinality / non_empty < 0.30:
        return "metadata", (
            f"avg {avg_words:.1f} words, cardinality {cardinality}/{non_empty} "
            f"= {cardinality / non_empty:.0%} (categorical)"
        )

    # Rule 4: half the values end with '?' → almost certainly the
    # question column. This is the most discriminative single signal
    # when headers are absent or generic.
    if q_mark_rate > 0.50:
        return "question", f"{q_mark_rate:.0%} of values end with '?'"

    # Rule 5: long prose → answer column.
    if avg_words >= 30:
        return "answer", f"avg {avg_words:.1f} words (long prose)"

    # Rule 6: medium-length text with no other signal → tentatively
    # context. The LLM frequently corrects this; that is fine — the
    # heuristic exists to provide a starting point, not the truth.
    return "context", f"avg {avg_words:.1f} words, no decisive signal"


def _sheet_q_mark_score(ws) -> tuple[int, int]:
    """Cheap signal of "how Q&A-shaped is this sheet?".

    Walk the first MAX_PROFILE_ROWS rows, every column, count the cells
    whose stripped text ends with '?'. The sheet with the most such
    cells is the one most likely to contain the RFI questions. Returns
    (question_mark_count, row_count) so callers can tie-break on size.
    """
    q_marks = 0
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    last_row = min(max_row, MAX_PROFILE_ROWS)
    for row_idx in range(1, last_row + 1):
        for col_idx in range(1, max_col + 1):
            v = _cell_to_str(ws.cell(row=row_idx, column=col_idx).value)
            if v and v.rstrip().endswith("?"):
                q_marks += 1
    return q_marks, max_row


def auto_detect_header_row(ws, scan_rows: int = 50) -> tuple[int, str]:
    """Pick the header row by two-pass scan, return (row, reason).

    ARCHITECTURAL DECISION: two passes, label-match first, '?'-density
    second. The two signals catch different file shapes:

    Pass 1 — exact header-token match.
      Walk rows 1..scan_rows. For each cell, check if its canonical
      form (lowercased, stripped, trailing ':' removed) is in
      QUESTION_HEADERS ∪ ANSWER_HEADERS. The first row containing
      ANY such match is the header. This is the right signal for
      clean Q&A files where a labelled header row exists somewhere
      in the file — the Guardian file has "Question" / "Answer" in
      row 2, the multi-sheet Futureproof file has them in row 7. The
      label is the strongest possible "this is the header" signal.

    Pass 2 — first long '?'-ending cell, use the row above.
      Only runs if pass 1 finds nothing. Walk rows 1..scan_rows for
      the first cell ending '?' with 5+ words (a real RFI question,
      not a form-field checkbox). Set header_row to that row - 1.
      Handles form-style files where the actual table has no labelled
      header — only the implicit "rows of questions" itself.

    Default — header_row = 1.
      If neither pass finds anything in the scan window. The safest
      fallback; for a file that fits neither shape, the human can
      still override with --header-row.

    Why "row above" in pass 2: the convention in label-headed files
    is "header at N, first question at N+1". Pass 1 already handles
    that case explicitly. In pass 2 (no label found), the first row
    of real questions IS the first data row, so the row above is the
    nominal header — even if it's blank or a section title, the
    column-role inference still works on row N+1 downwards.

    Why pass 1 takes priority: pass 2 would skip rows 2..N-1 of a
    file like Guardian where some questions don't end in '?' (e.g.
    "Describe your approach to X." with a full stop). Those rows
    contain real Q&A and must not be lost. The label-match signal is
    deterministic and lossless when a labelled header exists.
    """
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    last = min(max_row, scan_rows)

    # The token sets are intentionally narrow — exact-match only,
    # case-insensitive, trailing colon stripped. Multi-word headers
    # like "Gateway Questions" won't match, which is correct: only
    # cells whose value IS a header label (not contains one) should
    # mark a row as a header.
    #
    # We drop the single-letter aliases ("a", "q") at this scope: a
    # stray cell containing literal "A" or "Q" (e.g. a column letter
    # used as a row marker, a sequence indicator) is a frequent false
    # positive when scanning whole sheets. The aliases remain useful
    # in column-header context (where they map to a column ROLE) but
    # not here (where we are deciding which ROW is the header).
    header_tokens = (QUESTION_HEADERS | ANSWER_HEADERS) - {"a", "q"}

    # Pass 1: label-match.
    for row_idx in range(1, last + 1):
        for col_idx in range(1, max_col + 1):
            v = _cell_to_str(ws.cell(row=row_idx, column=col_idx).value)
            if not v:
                continue
            canonical = v.lower().strip().rstrip(":").strip()
            if canonical in header_tokens:
                return row_idx, (
                    f"row {row_idx} contains header label '{v}' in col "
                    f"{get_column_letter(col_idx)}"
                )

    # Pass 2: '?'-density fallback for files with no labelled header.
    for row_idx in range(1, last + 1):
        for col_idx in range(1, max_col + 1):
            v = _cell_to_str(ws.cell(row=row_idx, column=col_idx).value)
            if not v or not v.rstrip().endswith("?"):
                continue
            if len(v.split()) < 5:
                continue
            header_row = max(row_idx - 1, 1)
            return header_row, (
                f"no header label found; first long '?' question at "
                f"row {row_idx} (col {get_column_letter(col_idx)}, "
                f"{len(v.split())} words); header_row = {header_row}"
            )

    return 1, "no header signal in first 50 rows — default to header_row=1"


def pick_q_and_a_sheet(workbook) -> tuple[Any, str]:
    """Pick the sheet most likely to hold the Q&A data, and return why.

    ARCHITECTURAL DECISION: rank by question-mark density, fall back
    to row count. Earlier versions picked "largest sheet by row count"
    — that's wrong on files where the largest sheet is an operational
    plan / lookup table and the actual RFI sheet is smaller. Counting
    cells ending with '?' is the cheapest signal that correlates with
    "this sheet has questions in it". When no sheet has any question
    marks, the function falls back to the row-count rule (best
    available guess), and the human can still override with `--sheet`.
    """
    scored: list[tuple[int, int, Any]] = []  # (q_marks, max_row, ws)
    for name in workbook.sheetnames:
        ws = workbook[name]
        q_marks, max_row = _sheet_q_mark_score(ws)
        scored.append((q_marks, max_row, ws))
    # Sort: most question marks first, then most rows.
    scored.sort(key=lambda t: (-t[0], -t[1]))
    best = scored[0]
    if best[0] > 0:
        reason = f"highest question-mark count ({best[0]} cells ending '?')"
    else:
        reason = f"no '?'-ending cells in any sheet — picked largest by rows"
    return best[2], reason


# ─── Phase 2: LLM mapping ───────────────────────────────────────────────
def build_mapping_prompt(filename: str, sheet: SheetProfile) -> str:
    """Compose the user-message prompt for the role-mapping call.

    The prompt: (a) describes the task, (b) lists the allowed roles and
    constraints, (c) shows the column profile, (d) specifies the exact
    JSON output shape. We deliberately do NOT use a separate system
    message — Mistral small follows instructions reliably inside a
    single user message and it keeps the prompt easy to read end-to-end.
    """
    profile_lines: list[str] = []
    for c in sheet.columns:
        profile_lines.append(
            f"  Column {c.letter}: header='{c.header}', "
            f"non_empty={c.pct_non_empty:.0%}, "
            f"avg_words={c.avg_word_count:.1f}, "
            f"question_mark_rate={c.question_mark_rate:.0%}, "
            f"cardinality={c.cardinality}/{c.non_empty_count}, "
            f"samples={c.sample_values}, "
            f"heuristic_guess='{c.heuristic_role}' ({c.heuristic_reason})"
        )
    profile_block = "\n".join(profile_lines)

    return f"""You are profiling one sheet of an RFI Excel file to map its
columns to semantic roles. Reply with JSON only — no prose around it.

File: {filename}
Sheet: {sheet.name} ({sheet.n_data_rows} data rows, {sheet.n_cols} columns)

Column profile (the 'heuristic_guess' is from a simple rule-based
classifier — use it as a starting point, override when the stats suggest
otherwise):

{profile_block}

Map each column to ONE of these roles:
  - "question": the column containing RFI questions. EXACTLY ONE column.
  - "answer":   the column containing RFI answers. EXACTLY ONE column.
  - "context":  supporting context per row. ZERO OR ONE column.
  - "ignore":   columns with no semantic value (row numbers, blanks,
                internal admin). ZERO OR MORE columns.
  - any custom snake_case name (e.g. "category", "section", "ref"):
                metadata fields. ZERO OR MORE, one column per name.

Also infer best-effort (return null if you can't):
  - "client":   which client this RFI is for, from the filename or content.
  - "date":     year or date string, from the filename or content.

Return ONLY a JSON object with this shape (column letters as keys, no
trailing commas, no comments):

{{
  "column_roles": {{"A": "<role>", "B": "<role>", "..." : "..."}},
  "client": "<string or null>",
  "date": "<string or null>",
  "reasoning": "<1 to 3 sentences explaining the key decisions>"
}}
"""


def request_mapping(filename: str, sheet: SheetProfile) -> MappingProposal:
    """Call Mistral, parse JSON, return a MappingProposal.

    Every API call goes through call_with_retry. JSON-mode is requested
    via response_format; we still defensively parse and validate the
    response shape because mistralai 2.x is not guaranteed to return
    syntactically valid JSON in every case.
    """
    client = get_client()
    prompt = build_mapping_prompt(filename, sheet)
    response = call_with_retry(
        client.chat.complete,
        model=MISTRAL_MODEL,
        messages=[{"role": "user", "content": prompt}],
        response_format={"type": "json_object"},
        temperature=0.0,  # deterministic-ish; the profile is the same each run
    )
    text = response.choices[0].message.content
    try:
        payload = json.loads(text)
    except json.JSONDecodeError as exc:
        raise ValueError(
            f"Mistral returned text that is not valid JSON:\n{text}\n\n{exc}"
        ) from exc

    column_roles = payload.get("column_roles") or {}
    if not isinstance(column_roles, dict):
        raise ValueError("Mistral payload missing 'column_roles' object.")
    # Normalise: uppercase letters, stripped role strings.
    column_roles = {
        str(k).strip().upper(): str(v).strip().lower()
        for k, v in column_roles.items()
    }

    return MappingProposal(
        sheet=sheet.name,
        column_roles=column_roles,
        client=payload.get("client") or None,
        date=payload.get("date") or None,
        reasoning=str(payload.get("reasoning") or "").strip(),
    )


# ─── Phase 3: validation, display, approval, write ──────────────────────
def validate_proposal(proposal: MappingProposal, sheet: SheetProfile) -> list[str]:
    """Return a list of validation issues. Empty list = proposal is OK.

    We check three things the spec mandates:
      - every role is either reserved (question/answer/context/ignore)
        or a valid snake_case metadata name
      - exactly one column is `question` and exactly one is `answer`
      - every sheet column appears in the mapping (no silent drops)
    """
    issues: list[str] = []
    sheet_cols = {c.letter for c in sheet.columns}
    proposal_cols = set(proposal.column_roles.keys())

    missing = sheet_cols - proposal_cols
    extra = proposal_cols - sheet_cols
    if missing:
        issues.append(f"Mapping is missing roles for columns: {sorted(missing)}")
    if extra:
        issues.append(f"Mapping references unknown columns: {sorted(extra)}")

    role_counts: dict[str, int] = {}
    for col, role in proposal.column_roles.items():
        role_counts[role] = role_counts.get(role, 0) + 1
        if role in RESERVED_ROLES:
            continue
        if not ROLE_NAME_PATTERN.match(role):
            issues.append(
                f"Column {col}: role '{role}' is not a valid snake_case name "
                f"(must be lowercase, start with letter, [a-z0-9_])."
            )

    if role_counts.get("question", 0) != 1:
        issues.append(
            f"Expected exactly 1 'question' column, got {role_counts.get('question', 0)}."
        )
    if role_counts.get("answer", 0) != 1:
        issues.append(
            f"Expected exactly 1 'answer' column, got {role_counts.get('answer', 0)}."
        )
    if role_counts.get("context", 0) > 1:
        issues.append(
            f"Expected 0 or 1 'context' column, got {role_counts.get('context', 0)}."
        )

    # No duplicate metadata field names (two columns mapped to "category"
    # would silently collide in the config).
    for role, count in role_counts.items():
        if role in RESERVED_ROLES:
            continue
        if count > 1:
            issues.append(
                f"Metadata role '{role}' is assigned to {count} columns; "
                f"each metadata field must map to exactly one column."
            )

    return issues


def print_profile_table(sheet: SheetProfile) -> None:
    """Render the Phase 1 column profile as a readable table."""
    print(f"\n  Sheet: {sheet.name}  ({sheet.n_data_rows} data rows, "
          f"{sheet.n_cols} columns)")
    print("  " + "─" * 110)
    print(f"  {'Col':<4} {'Header':<24} {'NonE':>5} {'AvgW':>6} {'Q%':>5} "
          f"{'Card':>10}  Heuristic")
    print("  " + "─" * 110)
    for c in sheet.columns:
        card = f"{c.cardinality}/{c.non_empty_count}"
        print(f"  {c.letter:<4} {_truncate(c.header, 24):<24} "
              f"{c.pct_non_empty:>4.0%} {c.avg_word_count:>6.1f} "
              f"{c.question_mark_rate:>4.0%} {card:>10}  "
              f"{c.heuristic_role}")
        for s in c.sample_values:
            print(f"       └─ {_truncate(s, 90)}")


def print_proposal(proposal: MappingProposal, sheet: SheetProfile) -> None:
    """Render the validated Phase 2/3 proposal for the human."""
    print("\n  Proposed mapping (from Mistral, validated):")
    headers_by_letter = {c.letter: c.header for c in sheet.columns}
    for col in sorted(proposal.column_roles.keys(),
                      key=lambda x: (len(x), x)):
        role = proposal.column_roles[col]
        header = headers_by_letter.get(col, "?")
        marker = "(metadata)" if role not in RESERVED_ROLES else ""
        print(f"    {col}  '{_truncate(header, 30)}'  →  {role}  {marker}")
    print(f"\n  Inferred client: {proposal.client!r}")
    print(f"  Inferred date:   {proposal.date!r}")
    if proposal.reasoning:
        print(f"\n  Model reasoning: {proposal.reasoning}")


def prompt_yes_no(message: str) -> bool:
    """Read y/n from stdin. Defaults to NO on empty / ambiguous input.

    Defaulting to NO matches the spec's Decision 2: the approval gate is
    load-bearing. A user who just hits Enter has not approved. A user
    who types anything other than 'y' or 'yes' has not approved.
    """
    try:
        answer = input(f"\n  {message} [y/N] ").strip().lower()
    except EOFError:
        return False
    return answer in {"y", "yes"}


def slugify_for_config(source_path: Path) -> str:
    """Convert a filename to a safe config slug.

    "Utiq_Publicis RFI.xlsx"          → "utiq_publicis_rfi"
    "INTERNAL - Reach DPIA.xlsx"     → "internal_reach_dpia"
    Slugs are used in `config_rfi_<slug>.json` which is committed to git,
    so they must be ASCII and free of characters that need shell-quoting.
    """
    stem = source_path.stem.lower()
    slug = re.sub(r"[^a-z0-9]+", "_", stem)
    slug = re.sub(r"_+", "_", slug).strip("_")
    return slug or "untitled"


def build_config(
    proposal: MappingProposal,
    sheet: SheetProfile,
    source_path: Path,
    header_row: int,
) -> dict:
    """Convert the validated proposal into the on-disk config format
    described in the spec (Decision 2).

    Output shape:
      {
        "source_file": "<filename>",
        "sheet": "<sheet name>",
        "columns": {
          "question": "B",
          "answer": "C",
          "context": "D",          # only if a context column was mapped
          "<metadata_name>": "A",  # one entry per metadata column
          "ignore": ["E", "F"]      # always a list (possibly empty)
        },
        "metadata_fields": [...],   # custom role names from above
        "client": "...",
        "date": "..."
      }
    """
    columns: dict[str, Any] = {}
    metadata_fields: list[str] = []
    ignored: list[str] = []
    # Sort by (length, lexical) so columns appear in spreadsheet order:
    # A, B, ..., Z, AA, AB, ...
    for col in sorted(proposal.column_roles.keys(),
                      key=lambda x: (len(x), x)):
        role = proposal.column_roles[col]
        if role == "ignore":
            ignored.append(col)
        elif role in {"question", "answer", "context"}:
            columns[role] = col
        else:
            # Custom metadata role.
            columns[role] = col
            metadata_fields.append(role)
    columns["ignore"] = ignored

    # header_row is load-bearing for the loader: it determines which
    # spreadsheet row contains the column labels, and therefore which
    # row is the first data row (header_row + 1). Form-style RFIs have
    # header_row > 1 (e.g. 12 for Utiq_Publicis RFI). Without this
    # field the loader would default to 1 and silently misread the
    # file as starting at row 2 — answer cells loaded as questions,
    # preamble rows loaded as Q&A. Recording it in the config means
    # the loader does not re-run the discovery heuristic and the
    # mapping is reproducible.
    return {
        "source_file": source_path.name,
        "sheet": proposal.sheet,
        "header_row": header_row,
        "columns": columns,
        "metadata_fields": metadata_fields,
        "client": proposal.client,
        "date": proposal.date,
    }


def write_config(config: dict, path: Path) -> None:
    """Write the config as pretty-printed JSON. Overwrites with notice."""
    if path.exists():
        print(f"  Note: {path.name} already exists — overwriting.")
    path.write_text(json.dumps(config, indent=2, ensure_ascii=False) + "\n",
                    encoding="utf-8")
    print(f"  Wrote {path.name}.")


# ─── CLI ───────────────────────────────────────────────────────────────
def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Profile a single RFI Excel file and produce an "
                    "approved column→role mapping config."
    )
    parser.add_argument(
        "excel_path",
        type=Path,
        help="Path to the .xlsx file to profile (e.g. data/rfi_1.xlsx).",
    )
    parser.add_argument(
        "--sheet",
        type=str,
        default=None,
        help="Sheet name to profile. If omitted, auto-pick by "
             "question-mark density.",
    )
    parser.add_argument(
        "--header-row",
        type=int,
        default=None,
        dest="header_row",
        help="1-based row number of the column-header row. If omitted, "
             "the profiler auto-detects by scanning for the first row "
             "containing a real question (cell ending '?' with 5+ "
             "words). Use this flag to override when auto-detect picks "
             "wrong on a particular file.",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    source_path: Path = args.excel_path
    if not source_path.exists():
        print(f"ERROR: file not found: {source_path}", file=sys.stderr)
        return 2
    if source_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        print(f"ERROR: expected an .xlsx file, got {source_path.suffix}",
              file=sys.stderr)
        return 2

    print("=" * 76)
    print(f"RFI Excel Schema Profiler — {source_path}")
    print("=" * 76)

    # ── Phase 1
    print("\nPhase 1: Schema discovery")
    workbook = load_workbook(str(source_path), data_only=True, read_only=True)
    if not workbook.sheetnames:
        print("ERROR: workbook has no sheets.", file=sys.stderr)
        return 2

    print(f"  Sheets found: {len(workbook.sheetnames)}")
    for name in workbook.sheetnames:
        ws_each = workbook[name]
        print(f"    - {name}: rows={ws_each.max_row or 0}, "
              f"cols={ws_each.max_column or 0}")

    # Sheet selection: explicit --sheet wins; otherwise auto-pick by
    # question-mark density (see pick_q_and_a_sheet).
    if args.sheet is not None:
        if args.sheet not in workbook.sheetnames:
            print(f"\nERROR: --sheet '{args.sheet}' not found. Available: "
                  f"{workbook.sheetnames}", file=sys.stderr)
            return 2
        ws = workbook[args.sheet]
        print(f"  Sheet selected (explicit --sheet): {ws.title}")
    else:
        ws, reason = pick_q_and_a_sheet(workbook)
        print(f"  Sheet selected (auto): {ws.title}  [{reason}]")

    if args.header_row is not None:
        header_row = args.header_row
        print(f"  Header row (explicit --header-row): {header_row}")
    else:
        header_row, header_reason = auto_detect_header_row(ws)
        print(f"  Header row (auto): {header_row}  [{header_reason}]")

    try:
        sheet_profile = profile_sheet(ws, header_row=header_row)
    except ValueError as exc:
        print(f"\nERROR: {exc}", file=sys.stderr)
        return 2
    print_profile_table(sheet_profile)

    # ── Phase 2
    print("\nPhase 2: Mistral mapping recommendation")
    print(f"  Calling {MISTRAL_MODEL}...")
    try:
        proposal = request_mapping(source_path.name, sheet_profile)
    except Exception as exc:  # noqa: BLE001 — surface any LLM failure as a clean error
        print(f"\nERROR: Mistral mapping failed: {exc}", file=sys.stderr)
        return 3

    # ── Phase 3
    print("\nPhase 3: Validation and human approval")
    issues = validate_proposal(proposal, sheet_profile)
    if issues:
        print("\n  Proposal REJECTED by validator:")
        for issue in issues:
            print(f"    × {issue}")
        print("\n  Re-run the profiler; the LLM is non-deterministic across runs.")
        return 4

    print("  All checks passed.")
    print_proposal(proposal, sheet_profile)

    if not prompt_yes_no("Approve and write config?"):
        print("  Not approved — no config written.")
        return 1

    slug = slugify_for_config(source_path)
    config_path = Path(f"config_rfi_{slug}.json")
    config = build_config(proposal, sheet_profile, source_path,
                          header_row=header_row)
    write_config(config, config_path)
    return 0


if __name__ == "__main__":
    sys.exit(main())
