"""api.services.profiler — wrap pipeline.profile as an SSE stream.

Yields events in this order:

    {"type": "step",     "data": "<message>"}          (one per phase)
    {"type": "proposal", "data": {...full proposal...}}
    {"type": "done"}

If anything fails (invalid Excel, Mistral down, validation
rejects the LLM output), yields a single error event and stops:

    {"type": "error", "data": "<message>", "issues": [...]}

The profiler's CLI runs an interactive y/n approval at the end;
that approval is deliberately *not* in this service. The proposal
is written to `tmp/{session_id}/profile.json` and the human
approves via a separate POST endpoint (rfi_SPEC Step 3). Keeping
approval out of the SSE stream means a dropped connection between
"proposal yielded" and "approval clicked" does not lose work — the
proposal is on disk before the proposal event is sent.
"""

from __future__ import annotations

import asyncio
import json
from dataclasses import asdict
from pathlib import Path
from typing import AsyncGenerator

from openpyxl import load_workbook

from pipeline.profile import (
    auto_detect_header_row,
    pick_q_and_a_sheet,
    profile_sheet,
    request_mapping,
    validate_proposal,
)

UPLOAD_FILENAME = "upload.xlsx"
PROFILE_FILENAME = "profile.json"


# ARCHITECTURAL DECISION: every blocking pipeline call goes through
# asyncio.to_thread, the event loop never blocks.
#
# pipeline.profile.* are synchronous: openpyxl I/O, then a Mistral
# HTTP round-trip. Running them inline inside an async function
# blocks the FastAPI event loop and pauses every other in-flight
# SSE stream for the duration. asyncio.to_thread schedules them on
# the default thread executor — the event loop stays responsive,
# the synchronous functions need no rewrite, and concurrent
# sessions (multi-tab, multi-user) are isolated.
#
# This is the load-bearing reason the pipeline restructure (entry 15)
# stripped module-level side effects from pipeline.profile: thread
# execution requires the function be importable without firing
# argparse or opening a ChromaDB connection at import time.
async def _to_thread(func, *args, **kwargs):
    return await asyncio.to_thread(func, *args, **kwargs)


async def run_profile(session_dir: Path) -> AsyncGenerator[dict, None]:
    """Stream profiler events for the file at session_dir/upload.xlsx."""
    upload_path = session_dir / UPLOAD_FILENAME
    if not upload_path.exists():
        yield {"type": "error", "data": f"No upload found at {upload_path}"}
        return

    try:
        # Phase 1: open + sheet listing
        workbook = await _to_thread(
            load_workbook, str(upload_path), data_only=True, read_only=True
        )
        sheets = list(workbook.sheetnames)
        if not sheets:
            yield {"type": "error", "data": "Workbook has no sheets."}
            return
        yield {
            "type": "step",
            "data": f"File opened — {len(sheets)} sheet(s) detected",
        }

        # Phase 2: pick the Q&A sheet
        ws, pick_reason = await _to_thread(pick_q_and_a_sheet, workbook)
        yield {
            "type": "step",
            "data": f'Sheet selected: "{ws.title}" — {pick_reason}',
        }

        # Phase 3: detect header row
        header_row, header_reason = await _to_thread(auto_detect_header_row, ws)
        yield {
            "type": "step",
            "data": f"Header row: {header_row} — {header_reason}",
        }

        # Phase 4: per-column profile
        sheet_profile = await _to_thread(profile_sheet, ws, header_row)
        yield {
            "type": "step",
            "data": (
                f"Columns profiled: {sheet_profile.n_cols} columns, "
                f"{sheet_profile.n_data_rows} data rows"
            ),
        }

        # Phase 5: LLM mapping recommendation
        yield {
            "type": "step",
            "data": "Calling Mistral for column→role mapping...",
        }
        proposal = await _to_thread(
            request_mapping, upload_path.name, sheet_profile
        )
        yield {"type": "step", "data": "LLM recommendation received"}

        # Phase 6: validate
        issues = await _to_thread(validate_proposal, proposal, sheet_profile)
        if issues:
            yield {
                "type": "error",
                "data": "Proposal rejected by validator",
                "issues": issues,
            }
            return
        yield {"type": "step", "data": "Proposal validated"}

        # Phase 7: persist proposal, then yield it
        proposal_payload = {
            "source_file": upload_path.name,
            "sheet": proposal.sheet,
            "header_row": header_row,
            "column_roles": proposal.column_roles,
            "client": proposal.client,
            "date": proposal.date,
            "reasoning": proposal.reasoning,
            "columns": [
                {
                    "letter": c.letter,
                    "header": c.header,
                    "samples": c.sample_values,
                    "heuristic_role": c.heuristic_role,
                }
                for c in sheet_profile.columns
            ],
        }
        # ARCHITECTURAL DECISION: write proposal.json BEFORE yielding the
        # proposal event. If the SSE connection drops between the yield
        # and the human clicking "approve", the proposal still survives
        # on disk and the approval POST in Step 3 can pick it up.
        # Yielding first and writing after would create a window where
        # the user sees the proposal but the backend has no record of it.
        (session_dir / PROFILE_FILENAME).write_text(
            json.dumps(proposal_payload, indent=2, ensure_ascii=False) + "\n",
            encoding="utf-8",
        )
        yield {"type": "proposal", "data": proposal_payload}
        yield {"type": "done"}

    except Exception as exc:  # noqa: BLE001 — surface all failures as an error event
        # ARCHITECTURAL DECISION: catch-all → error event, not bubbled exception.
        # An exception thrown inside an async generator that is the source of a
        # StreamingResponse becomes an opaque 500 on the wire — the SSE client
        # sees the stream close with no useful message. Catching here and
        # yielding a typed error event preserves the per-event protocol all
        # the way to the browser, where the frontend can render it.
        yield {"type": "error", "data": f"{type(exc).__name__}: {exc}"}
