"""api.services.answerer — wrap pipeline.query as an SSE stream.

The answer workflow: for each question in a freshly-uploaded RFI,
retrieve relevant past Q&A from the corpus, rerank, generate an
answer, and yield an event with the answer + full retrieval trace.

Events yielded:

    {"type": "progress", "data": {"index": 1, "total": 47,
                                    "question_text": "..."}}
    {"type": "answer",   "data": {
        "index": 1,
        "question": "...",
        "answer":   "...",
        "refused":  false,
        "confidence": 0.91,
        "sources": [
            {"rank": 1, "source_file": "...", "pair_id": "...",
             "section": "...", "score": 0.91,
             "score_type": "crossencoder",
             "question_text": "...", "answer_text": "..."},
            ...
        ],
        "pair_ids": ["..."],
        "mentioned_clients": ["Reach"]   # cross-tenant leakage flag
    }}
    {"type": "done", "data": {"answered": 45, "refused": 2, "total": 47}}

Errors yield {"type": "error", "data": "..."} and stop the stream.

ARCHITECTURAL DECISION: production config from LEARNING_NOTES
entry 13, not the rfi_SPEC Step 4 prompt's older recommendation.

rfi_SPEC was written before the eval finalised; its Step 4 prompt
specified `rfi_separated_cosine + hybrid + crossencoder + top-k=3`.
Entry 13's eval determined that semantic retrieval beat hybrid on
this small/paraphrase-rich corpus, so the production
recommendation is semantic (not hybrid). We follow entry 13 — it's
the empirical winner against the ground-truth question set, and
making the UI default match the eval's recommendation is what
keeps the eval load-bearing rather than ceremonial.

ARCHITECTURAL DECISION: surface cross-tenant client mentions per
answer.

LEARNING_NOTES entry 14 documents that generated answers can name
past clients verbatim because the retrieved chunks do. The fix
(prompt guard + post-redaction) is not implemented at the pipeline
layer yet. Until it is, the UI's job is to flag the risk so the
human reviewer can catch it before sending to the new client. We
list every known client name (collected from `config_rfi_*.json`
files at repo root) that appears in the generated answer text,
under `mentioned_clients`. The frontend renders this as a visible
warning badge per AnswerCard.

This is "do not ship a 'send directly to client' path" expressed
as code: every answer that names a past client is marked, the
reviewer sees the mark before approving.
"""

from __future__ import annotations

import asyncio
import glob
import json
import re
from pathlib import Path
from typing import AsyncGenerator

from openpyxl import load_workbook

from api.chroma_client import get_chroma_client
from pipeline.mistral_helpers import get_client
from pipeline.profile import (
    auto_detect_header_row,
    pick_q_and_a_sheet,
    profile_sheet,
    request_mapping,
)
from pipeline.query import (
    DEFAULT_POOL_SIZE,
    DEFAULT_TOP_K,
    ChunkResult,
    fetch_paired_answers,
    generate_answer,
    rerank_crossencoder,
    retrieve_semantic,
)

# Production config — see LEARNING_NOTES entry 13.
DEFAULT_COLLECTION = "rfi_separated_cosine"
DEFAULT_RETRIEVAL = "semantic"
DEFAULT_RERANK = "crossencoder"
REFUSAL_TEXT = "I cannot find this in our corpus."

QUESTIONS_FILENAME = "answer_questions.json"
ANSWERS_FILENAME = "answers.json"
UPLOAD_FILENAME = "upload.xlsx"


# ARCHITECTURAL DECISION: heuristic first, LLM fallback for the
# question column.
#
# The lightweight goal is "find the question column without making
# the user wait". The heuristic (pipeline.profile.profile_sheet's
# heuristic_role) catches the easy case in milliseconds: header
# matches "Question" or >50% of cells end with "?". But it misses
# files where the question column is real prose with no "?"
# (e.g. "Please describe your approach to X.") and the header is
# a topic name like "Company Overview". The ingest workflow handles
# this with a full Mistral call; the answer workflow needs the
# same robustness.
#
# Solution: try the heuristic first; if it returns zero candidates
# OR more than one with no clear winner, fall back to a Mistral
# mapping call (same `request_mapping` the ingest profiler uses)
# and read the column with `role == "question"`. The fast path
# stays fast for files that fit the heuristic; the slow path costs
# one Mistral round-trip (~1-3s) for the harder files. The user
# sees which method was used in the upload response, so the
# question-column decision is auditable on every run.
def extract_questions(upload_path: Path) -> dict:
    """Open the uploaded Excel, find the question column, return
    {sheet, header_row, question_column, questions: [{row, text}]}.

    Tries the heuristic from pipeline.profile.profile_sheet first;
    falls back to a Mistral mapping call when the heuristic cannot
    confidently pick a single column.

    Raises ValueError if neither method identifies a question
    column.
    """
    workbook = load_workbook(str(upload_path), data_only=True, read_only=True)
    if not workbook.sheetnames:
        raise ValueError("Workbook has no sheets.")

    ws, sheet_reason = pick_q_and_a_sheet(workbook)
    header_row, header_reason = auto_detect_header_row(ws)
    sheet_profile = profile_sheet(ws, header_row=header_row)

    candidates = [
        c for c in sheet_profile.columns
        if c.heuristic_role == "question"
    ]
    detection_method: str
    if len(candidates) == 1:
        best = candidates[0]
        detection_method = "heuristic"
    elif len(candidates) > 1:
        # Multiple heuristic guesses — tie-break by q_mark_rate.
        best = max(candidates, key=lambda c: c.question_mark_rate)
        detection_method = "heuristic (tie-broken on q_mark_rate)"
    else:
        # Fall back to Mistral.
        proposal = request_mapping(upload_path.name, sheet_profile)
        question_col = next(
            (col for col, role in proposal.column_roles.items()
             if role == "question"),
            None,
        )
        if not question_col:
            raise ValueError(
                "Neither the heuristic nor Mistral identified a "
                "question column. The file may not be a Q&A RFI. "
                f"Sheet auto-pick was: {sheet_reason}"
            )
        best_match = next(
            (c for c in sheet_profile.columns if c.letter == question_col),
            None,
        )
        if best_match is None:
            raise ValueError(
                f"Mistral named column {question_col!r} as 'question' but "
                f"that column is not in the profiled sheet."
            )
        best = best_match
        detection_method = "llm-fallback"

    # Walk the column from header_row+1 to end, collect non-empty cells.
    question_col_letter = best.letter
    col_idx = ord(question_col_letter.upper()) - ord("A") + 1
    # Letters past Z aren't expected for RFI files but handle just in case.
    if question_col_letter and len(question_col_letter) > 1:
        col_idx = 0
        for ch in question_col_letter.upper():
            col_idx = col_idx * 26 + (ord(ch) - ord("A") + 1)

    questions: list[dict] = []
    max_row = ws.max_row or 0
    for r in range(header_row + 1, max_row + 1):
        val = ws.cell(row=r, column=col_idx).value
        if val is None:
            continue
        text = str(val).strip()
        if not text:
            continue
        questions.append({"row": r, "text": text})

    return {
        "sheet": ws.title,
        "sheet_reason": sheet_reason,
        "header_row": header_row,
        "header_reason": header_reason,
        "question_column": question_col_letter,
        "question_column_header": best.header,
        "detection_method": detection_method,
        "questions": questions,
    }


# ── Per-answer cross-tenant leakage check ─────────────────────────────
def _load_known_clients() -> list[str]:
    """Collect client names from every config_rfi_*.json at repo root.

    These are the clients the corpus draws from — any one of them
    appearing in a generated answer is a leakage candidate that the
    human reviewer should see before accepting the answer.
    """
    names: set[str] = set()
    for path in glob.glob("config_rfi_*.json"):
        try:
            with open(path, encoding="utf-8") as f:
                cfg = json.load(f)
            client = cfg.get("client")
            if isinstance(client, str) and client.strip():
                names.add(client.strip())
        except (OSError, json.JSONDecodeError):
            continue
    return sorted(names)


def _mentions_of_clients(answer_text: str, clients: list[str]) -> list[str]:
    """Return the subset of `clients` whose names appear in `answer_text`
    as a whole word (case-insensitive)."""
    lowered = answer_text.lower()
    hits: list[str] = []
    for name in clients:
        # Word-boundary match to avoid "Reach" matching "outreach", etc.
        pattern = r"\b" + re.escape(name.lower()) + r"\b"
        if re.search(pattern, lowered):
            hits.append(name)
    return hits


# ── Per-question retrieve + rerank + generate ─────────────────────────
def _build_answer_payload(
    index: int,
    question: str,
    answer_text: str,
    top: list[ChunkResult],
    paired: list[dict],
    known_clients: list[str],
) -> dict:
    """Compose the SSE answer event's `data` block with full provenance."""
    refused = answer_text.strip() == REFUSAL_TEXT
    confidence = top[0].score if top else 0.0

    sources: list[dict] = []
    for rank, (q_chunk, a) in enumerate(zip(top, paired), 1):
        sources.append({
            "rank": rank,
            "source_file": q_chunk.metadata.get("source_file", "?"),
            "pair_id": q_chunk.metadata.get("pair_id", "?"),
            "section": q_chunk.metadata.get("section"),
            "client": q_chunk.metadata.get("client"),
            "score": round(float(q_chunk.score), 4),
            "score_type": q_chunk.score_type,
            "question_text": q_chunk.text,
            "answer_text": a.get("text", ""),
        })

    mentioned = _mentions_of_clients(answer_text, known_clients)
    return {
        "index": index,
        "question": question,
        "answer": answer_text,
        "refused": refused,
        "confidence": round(float(confidence), 4),
        "sources": sources,
        "pair_ids": [q.metadata.get("pair_id", "?") for q in top],
        "mentioned_clients": mentioned,
    }


def _answer_one_question(
    chroma_collection,
    mistral_client,
    question: str,
    known_clients: list[str],
    index: int,
) -> dict:
    """Synchronous helper: retrieve + rerank + generate for one question.

    Called via asyncio.to_thread by the generator below so the event
    loop stays unblocked. Returns the answer payload dict ready to
    yield.
    """
    # Retrieve question chunks only — separated collection.
    pool = retrieve_semantic(
        chroma_collection,
        mistral_client,
        question,
        n=DEFAULT_POOL_SIZE,
        where={"role": "question"},
    )
    if not pool:
        # Empty corpus or pathological query — refusal is correct.
        return _build_answer_payload(
            index=index, question=question, answer_text=REFUSAL_TEXT,
            top=[], paired=[], known_clients=known_clients,
        )

    top = rerank_crossencoder(question, pool, k=DEFAULT_TOP_K)
    paired = fetch_paired_answers(chroma_collection, top)
    answer_text = generate_answer(mistral_client, question, top, paired)
    return _build_answer_payload(
        index=index, question=question, answer_text=answer_text,
        top=top, paired=paired, known_clients=known_clients,
    )


# ── Top-level async generator ─────────────────────────────────────────
async def run_answer(session_dir: Path) -> AsyncGenerator[dict, None]:
    """Stream per-question answer events for the session's uploaded RFI."""
    questions_path = session_dir / QUESTIONS_FILENAME
    if not questions_path.exists():
        yield {
            "type": "error",
            "data": f"No {QUESTIONS_FILENAME} — run POST /api/answer/upload first",
        }
        return

    try:
        extracted = json.loads(questions_path.read_text(encoding="utf-8"))
        questions: list[dict] = extracted.get("questions", [])
        if not questions:
            yield {"type": "error", "data": "No questions found in the upload."}
            return

        chroma_client = await asyncio.to_thread(get_chroma_client)
        try:
            collection = await asyncio.to_thread(
                chroma_client.get_collection, DEFAULT_COLLECTION
            )
        except Exception as exc:  # noqa: BLE001
            yield {
                "type": "error",
                "data": (
                    f"Collection {DEFAULT_COLLECTION!r} not found ({exc}). "
                    f"Ingest at least one RFI first."
                ),
            }
            return

        mistral_client = get_client()
        known_clients = _load_known_clients()
        total = len(questions)
        answers: list[dict] = []
        refused_count = 0

        for i, q in enumerate(questions, 1):
            question_text = q["text"]
            yield {
                "type": "progress",
                "data": {
                    "index": i,
                    "total": total,
                    "question_text": question_text,
                },
            }
            payload = await asyncio.to_thread(
                _answer_one_question,
                collection,
                mistral_client,
                question_text,
                known_clients,
                i,
            )
            # Preserve the original row number on disk so the exporter
            # (Step 5) can write answers back into the correct rows.
            payload["row"] = q.get("row")
            answers.append(payload)
            if payload["refused"]:
                refused_count += 1
            yield {"type": "answer", "data": payload}

        # Persist the full answers list for the exporter.
        (session_dir / ANSWERS_FILENAME).write_text(
            json.dumps(answers, indent=2, ensure_ascii=False) + "\n",
            encoding="utf-8",
        )
        yield {
            "type": "done",
            "data": {
                "answered": total - refused_count,
                "refused": refused_count,
                "total": total,
            },
        }

    except Exception as exc:  # noqa: BLE001
        yield {"type": "error", "data": f"{type(exc).__name__}: {exc}"}
