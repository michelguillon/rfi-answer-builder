"""api.services.exporter — build the filled-in Excel for download.

Reads tmp/{sid}/upload.xlsx and tmp/{sid}/answers.json, appends
three columns to the data sheet (Suggested Answer, Source RFIs,
Confidence) and writes tmp/{sid}/output.xlsx. The router returns
that file as a FileResponse so the browser downloads it.

ARCHITECTURAL DECISION: open the workbook with data_only=False
and write back in place.

The profiler/upload step uses `data_only=True` to read formula
*results* for column-role classification. The exporter does NOT
read formula cells — it only appends NEW columns past the last
existing one and writes their header + per-row text. Opening with
`data_only=False` (the openpyxl default) preserves formulas,
formatting, merged cells, validation rules, and conditional
formatting in the existing columns. We do not touch them.

What openpyxl cannot preserve through a round-trip: VBA macros
(.xlsm files re-saved as .xlsx drop them), embedded charts in
some shapes, ActiveX controls. RFI files are typically plain Q&A
tables so this is a non-issue in practice. If a future RFI ships
with critical macros, the workaround is to instruct the user to
copy the new columns into the original file by hand rather than
losing the macros.

ARCHITECTURAL DECISION: refusal text lands in the cell, not
blank.

When the generator refuses ("I cannot find this in our corpus."),
that *is* the answer the system produced — it's load-bearing
information for the reviewer ("we don't have anything to say
here, you'll need to draft this manually"). Leaving the cell
blank would hide the refusal and make it look like the system
missed the row. Writing the refusal text into Suggested Answer
keeps the audit trail intact. The reviewer can clear the cell
manually if they want a blank export.

Skipped answers (the user actively rejected the suggestion in
the UI) DO get a blank cell — the user has chosen not to use the
suggestion at all.
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

PROFILE_FILENAME = "answer_questions.json"
ANSWERS_FILENAME = "answers.json"
UPLOAD_FILENAME = "upload.xlsx"
OUTPUT_FILENAME = "output.xlsx"
ORIGINAL_NAME_FILENAME = "original_filename"

HEADER_SUGGESTED = "Suggested Answer"
HEADER_SOURCES = "Source RFIs"
HEADER_CONFIDENCE = "Confidence"

_ROW_FROM_PAIR_ID = re.compile(r"_row_(\d+)$")


def _row_from_pair_id(pair_id: str) -> str:
    """Extract the Excel row number from a pair_id of the form
    "<slug>_row_<N>". Returns the row as a string for display, or
    the full pair_id if the regex misses (defensive — keeps the
    cell content useful rather than silently dropping the
    attribution).
    """
    if not pair_id:
        return "?"
    m = _ROW_FROM_PAIR_ID.search(pair_id)
    return m.group(1) if m else pair_id


def _format_sources(sources: list[dict]) -> str:
    """Build the "Source RFIs" cell value: pipe-separated list of
    "<source_file> row <N>" entries, one per retrieved chunk.

    Preserves the rank order — the most relevant source is first.
    Empty list -> empty string (skipped or empty-corpus refusal).
    """
    if not sources:
        return ""
    pieces: list[str] = []
    for s in sources:
        src = s.get("source_file") or "?"
        row = _row_from_pair_id(s.get("pair_id", ""))
        pieces.append(f"{src} row {row}")
    return " | ".join(pieces)


def build_output(session_dir: Path) -> Path:
    """Build tmp/{sid}/output.xlsx from upload.xlsx + answers.json.

    Each entry in answers.json may carry a `_status` field set by
    /api/answer/edit:
      - "accepted" (or absent): write generated answer
      - "edited":               `answer` field already replaced;
                                write as-is
      - "skipped":              write blank cells

    Returns the path to the written file. Raises FileNotFoundError /
    ValueError on missing inputs.
    """
    upload_path = session_dir / UPLOAD_FILENAME
    answers_path = session_dir / ANSWERS_FILENAME
    profile_path = session_dir / PROFILE_FILENAME

    if not upload_path.exists():
        raise FileNotFoundError(f"No upload at {upload_path}")
    if not answers_path.exists():
        raise FileNotFoundError(
            f"No answers.json at {answers_path}. Run the process step first."
        )
    if not profile_path.exists():
        raise FileNotFoundError(
            f"No {PROFILE_FILENAME} at {profile_path}. Re-upload."
        )

    answers: list[dict] = json.loads(answers_path.read_text(encoding="utf-8"))
    profile: dict[str, Any] = json.loads(profile_path.read_text(encoding="utf-8"))

    sheet_name: str = profile["sheet"]
    header_row: int = int(profile["header_row"])

    # Open with full fidelity — preserves formulas, formatting, etc.
    workbook = load_workbook(str(upload_path))
    if sheet_name not in workbook.sheetnames:
        raise ValueError(
            f"Sheet {sheet_name!r} not in workbook (have: {workbook.sheetnames}). "
            f"The upload may differ from what was profiled."
        )
    ws = workbook[sheet_name]

    # Append the three new columns past the last used column. We DO
    # NOT rely on the original profile's column count — the user
    # might have a workbook where the data sheet has trailing
    # decorative columns; we want our headers strictly after
    # everything currently in use.
    first_new_col = (ws.max_column or 0) + 1
    col_suggested = first_new_col
    col_sources = first_new_col + 1
    col_confidence = first_new_col + 2

    ws.cell(row=header_row, column=col_suggested, value=HEADER_SUGGESTED)
    ws.cell(row=header_row, column=col_sources, value=HEADER_SOURCES)
    ws.cell(row=header_row, column=col_confidence, value=HEADER_CONFIDENCE)

    for a in answers:
        row = a.get("row")
        if not isinstance(row, int):
            continue  # answer without a row (e.g. extracted from a non-row source)
        status = a.get("_status", "accepted")
        if status == "skipped":
            ws.cell(row=row, column=col_suggested, value=None)
            ws.cell(row=row, column=col_sources, value=None)
            ws.cell(row=row, column=col_confidence, value=None)
            continue
        ws.cell(row=row, column=col_suggested, value=a.get("answer") or "")
        ws.cell(row=row, column=col_sources, value=_format_sources(a.get("sources") or []))
        conf = a.get("confidence")
        if isinstance(conf, (int, float)):
            ws.cell(row=row, column=col_confidence, value=round(float(conf), 2))
        else:
            ws.cell(row=row, column=col_confidence, value=None)

    output_path = session_dir / OUTPUT_FILENAME
    workbook.save(str(output_path))
    return output_path


def get_download_name(session_dir: Path) -> str:
    """Read the original filename sidecar and append "_answered" to
    the stem. Falls back to "output.xlsx" if the sidecar is missing.
    """
    sidecar = session_dir / ORIGINAL_NAME_FILENAME
    if not sidecar.exists():
        return OUTPUT_FILENAME
    original = sidecar.read_text(encoding="utf-8").strip()
    if not original:
        return OUTPUT_FILENAME
    stem = Path(original).stem
    return f"{stem}_answered.xlsx"


def update_answers_inplace(session_dir: Path,
                           overrides: dict[int, str],
                           skipped: list[int]) -> int:
    """Persist user edits back into tmp/{sid}/answers.json.

    The frontend POSTs edits via /api/answer/edit before triggering
    the GET /api/answer/export download. We persist them so the
    download URL stays a simple GET (browser-friendly) — the export
    reads the already-edited answers.json straight off disk.

    Returns the number of records modified.
    """
    answers_path = session_dir / ANSWERS_FILENAME
    if not answers_path.exists():
        raise FileNotFoundError(f"No answers.json at {answers_path}")
    answers: list[dict] = json.loads(answers_path.read_text(encoding="utf-8"))
    skipped_set = set(skipped)
    modified = 0
    for a in answers:
        idx = a.get("index")
        if idx in skipped_set:
            a["_status"] = "skipped"
            modified += 1
        elif idx in overrides:
            a["answer"] = overrides[idx]
            a["refused"] = False
            a["_status"] = "edited"
            modified += 1
        elif "_status" not in a:
            a["_status"] = "accepted"
    answers_path.write_text(
        json.dumps(answers, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    return modified
