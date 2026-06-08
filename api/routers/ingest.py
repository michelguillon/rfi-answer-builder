"""api.routers.ingest — upload + profile (SSE) + approve (SSE).

  POST /api/ingest/upload?session_id=...     save Excel under tmp/{sid}/
  GET  /api/ingest/profile?session_id=...    SSE stream of profiler steps
  POST /api/ingest/approve                   SSE stream of ingest progress
"""

from __future__ import annotations

import asyncio
import json

from fastapi import APIRouter, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from pydantic import BaseModel

from api.services.ingester import ORIGINAL_NAME_FILENAME, run_ingest
from api.services.profiler import UPLOAD_FILENAME, run_profile
from api.session import get_session_dir

router = APIRouter(prefix="/api/ingest", tags=["ingest"])

ALLOWED_EXTENSIONS = {".xlsx", ".xlsm"}


# ARCHITECTURAL DECISION: upload saves to a fixed filename
# (tmp/{sid}/upload.xlsx), not the original name.
#
# Three reasons:
#   1. The session directory is the only state location; the profiler
#      service does not need to be told *which* file in the directory
#      to open. A fixed name removes that contract from the API.
#   2. Each session holds one workflow; re-uploading replaces. That
#      matches the spec's wizard shape (Step 1 = single upload).
#   3. Original filenames carry client names (e.g.
#      "Utiq_Publicis RFI.xlsx"). Storing them at a stable session-
#      scoped path keeps that PII out of any future filesystem listing
#      that might leak in error messages.
#
# The original filename is still preserved in the response payload
# so the frontend can display it back to the user — we just don't
# use it as the on-disk name.
@router.post("/upload")
async def upload(
    file: UploadFile = File(...),
    session_id: str = Query(...),
) -> dict:
    """Save the uploaded Excel to tmp/{session_id}/upload.xlsx.

    Returns {session_id, filename, detected_rows}.
    """
    if file.filename is None:
        raise HTTPException(400, "Missing filename")
    suffix = "." + file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if suffix not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            400,
            f"Only {sorted(ALLOWED_EXTENSIONS)} files accepted, got '{file.filename}'",
        )

    session_dir = get_session_dir(session_id)
    upload_path = session_dir / UPLOAD_FILENAME
    contents = await file.read()
    upload_path.write_bytes(contents)

    # ARCHITECTURAL DECISION: persist the original filename as a
    # sidecar in the session directory. The ingester (Step 3) needs
    # it later to write config_rfi_<slug>.json + copy upload.xlsx
    # into data/<original_filename>.xlsx so a future CLI re-ingest
    # works. The upload response already returns the filename to
    # the client, but storing it server-side keeps the contract one-
    # sided: the backend remembers, the frontend just kicks the
    # workflow.
    (session_dir / ORIGINAL_NAME_FILENAME).write_text(
        file.filename, encoding="utf-8"
    )

    detected_rows = await asyncio.to_thread(_estimate_rows, upload_path)
    return {
        "session_id": session_id,
        "filename": file.filename,
        "detected_rows": detected_rows,
    }


def _estimate_rows(path) -> int:
    """Row count of the largest sheet. Cheap; used only for upload preview."""
    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        return max((ws.max_row or 0) for ws in wb.worksheets) if wb.worksheets else 0
    finally:
        wb.close()


# ARCHITECTURAL DECISION: SSE response headers are explicit
# (Cache-Control + X-Accel-Buffering).
#
# A reverse proxy that buffers the response defeats the entire
# point of SSE — the client would see all events arrive at once on
# stream close instead of incrementally. The two headers below tell
# nginx (`X-Accel-Buffering: no`) and any HTTP cache
# (`Cache-Control: no-cache`) not to coalesce. They are listed in
# api/CLAUDE.md as the canonical SSE response shape; reuse for every
# SSE endpoint in this app.
SSE_HEADERS = {
    "Cache-Control": "no-cache",
    "X-Accel-Buffering": "no",
}


@router.get("/profile")
async def profile(session_id: str = Query(...)) -> StreamingResponse:
    """Stream profiler events as Server-Sent Events.

    Yields step / proposal / done events. On error yields a single
    error event and closes the stream.
    """
    session_dir = get_session_dir(session_id)

    async def stream():
        async for event in run_profile(session_dir):
            yield f"data: {json.dumps(event)}\n\n"

    return StreamingResponse(
        stream(),
        media_type="text/event-stream",
        headers=SSE_HEADERS,
    )


# ARCHITECTURAL DECISION: approve POSTs an SSE response.
#
# The spec calls for the approve endpoint to BOTH commit the user's
# config edits AND stream the ingest progress, so we return a
# StreamingResponse directly from a POST handler. The browser's
# native EventSource API only supports GET, so the frontend uses
# the fetch + ReadableStream pattern (documented in api/CLAUDE.md
# and the frontend SSE wrapper). This shape keeps the workflow at
# one request rather than splitting into "POST commit + GET stream"
# — fewer round trips and no risk of an interleaving race where
# the GET begins before the commit lands.
class ApproveBody(BaseModel):
    """Body of POST /api/ingest/approve.

    The frontend collects only client/date edits — column roles are
    NOT editable in the UI (see rfi_SPEC's Ingest wireframe). If
    omitted, the inferred values from profile.json are kept.
    """

    session_id: str
    client: str | None = None
    date: str | None = None


@router.post("/approve")
async def approve(body: ApproveBody) -> StreamingResponse:
    """Write the approved config and stream ingest progress."""
    session_dir = get_session_dir(body.session_id)

    async def stream():
        async for event in run_ingest(session_dir, body.client, body.date):
            yield f"data: {json.dumps(event)}\n\n"

    return StreamingResponse(
        stream(),
        media_type="text/event-stream",
        headers=SSE_HEADERS,
    )
