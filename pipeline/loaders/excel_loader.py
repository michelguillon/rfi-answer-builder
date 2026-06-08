"""
loaders/excel_loader.py — RFI Excel loader  [spec Step 3]
==========================================================
Reads one RFI Excel file using an approved config (produced by
`profile_excel.py`) and produces a list of `Row` dataclasses ready
for the chunker.

ARCHITECTURAL DECISION: separate signature from the prose loaders.
The docx and pdf loaders are `load_xxx(path) -> list[Paragraph]`.
The Excel loader is `load_excel(path, config) -> list[Row]`. The
asymmetry is honest — prose loaders discover everything they need
from the file format alone; an Excel file requires a human-approved
column→role mapping that cannot be inferred at load time. Forcing
both into a shared dispatcher (`load(path)`) would either smuggle
hidden state into the prose path or pretend Excel can self-configure.
Two functions, two contracts, one of which takes a config.

ARCHITECTURAL DECISION: pair_id uses the Excel row number, not a
sequential list index. pair_id is the linkage between a question
chunk and its paired answer chunk in the separated chunking strategy.
Two identifier schemes were considered:
  - Excel row number (1-based, preserves the original cell position
    even when intermediate rows are skipped).
  - List index (0-based, dense after skipping).
Excel-row wins on traceability — a retrieved chunk attributes back
to the exact spreadsheet row, which matters when the human reading
an answer wants to open the source file and verify it. The cost
(non-dense IDs) is paid in exchange for that traceability. A re-run
of the loader on the same file produces the same pair_ids, which
is what we want for deterministic Q↔A linkage downstream.

ARCHITECTURAL DECISION: skip a row when the QUESTION cell is empty.
This rule was tightened from "skip only when BOTH question and
answer are empty" after running the loader on real data. The
empirical motivation: the original rule kept rows where an answer
existed but the question column was blank, on the theory that an
answer alone might carry useful content. In practice, every such
row in the corpus turned out to be a profiler mis-mapping (most
visibly on `INTERNAL - Reach Customer facing DPIA questions.xlsx`,
where the LLM tagged the "Gateway Questions" column as `context`
and the sub-detail column as `question`, producing rows of the
form Q='', A='no', C='<the real question>'). Those rows can't be
retrieved by question similarity (no question to match against),
so admitting them costs corpus quality with no retrieval upside.
The tighter rule treats a row with no question as no Q&A row, full
stop. Rows where the question is present but the answer is blank
("asked but unanswered") are still loaded — that case is real and
preserved by the rule.

ARCHITECTURAL DECISION: metadata is built per-row at load time, not
deferred to the chunker. Every Row carries its own metadata dict
containing (a) row-level metadata columns (category, ref, etc., as
defined in config['metadata_fields']) and (b) file-level client and
date from the config top level. The chunker therefore needs no
knowledge of which file a Row came from to attach client/date — the
Row is self-contained. Trade-off: every Row from the same file
carries redundant client/date strings. Acceptable: row counts are
in the low thousands and the simplicity of "the Row knows everything
about itself" is worth the few extra bytes per row.

ARCHITECTURAL DECISION: read with iter_rows(values_only=True) rather
than ws.cell(). openpyxl in read_only mode optimises for streaming;
calling ws.cell(row, col) for every cell can degrade to O(n²) on
large sheets because each call may re-parse the row. iter_rows yields
one tuple per row, in order, with cell values pre-extracted. For an
88-column file we read ~5 of those values per row and discard the
rest, which is the right shape for sparse-column Excel layouts.

ARCHITECTURAL DECISION: detect section-divider rows and propagate
their value as metadata, do not emit them as Q&A pairs.
Empirically (see the diagnostic table in docs/rfi_LEARNING_NOTES.md
entry 6), three of the four real RFI files contain "section
marker" rows: a short label like "Audiences/Targeting" or "Direct
consent" sitting in the question column above the actual Q&A rows
for that section, with the answer cell empty. Loading these as Q&A
pairs would corrupt the corpus with non-question chunks (a chunk
whose text is just "Audiences/Targeting" has no semantic value and
pollutes retrieval).

The detection rule is intentionally conservative:
  - question cell has 1..4 words
  - question cell does NOT end with '?'
  - answer cell is empty
  - context cell (if present in config) is empty
  - every configured metadata cell is empty
A row matching all five conditions is a section marker. Any one of
them failing means we treat the row as a normal Q&A row. This loses
no legitimate Q&A (a real question may be short, but if it has any
answer/context/metadata it falls through and gets loaded) and keeps
the heuristic stupid enough to debug.

When the file has a dedicated section column (file 4 / Guardian:
the LLM identified a `section` metadata column), it produces
metadata['section']=<value> on every row directly. Inferred
section-marker propagation defers to that — if the metadata already
contains 'section', we don't overwrite it.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from pipeline.models import Row


# ─── Helpers ────────────────────────────────────────────────────────────
def _cell_str(value: Any) -> str:
    """Coerce an openpyxl cell value to a clean string. None becomes ''.

    openpyxl returns native Python types — int, float, datetime, str,
    None. For Row purposes we want strings; the loader is upstream of
    chunking and embedding, both of which want text.
    """
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _cell_at(row_values: tuple, col_1based: int) -> str:
    """Safely fetch a cell value by 1-based column index from a tuple
    returned by iter_rows(values_only=True). Out-of-range or None → ''."""
    if col_1based <= 0 or col_1based > len(row_values):
        return ""
    return _cell_str(row_values[col_1based - 1])


def _slug(stem: str) -> str:
    """Same slug rule as profile_excel.py — pair_ids must be readable
    and filesystem-safe."""
    s = re.sub(r"[^a-z0-9]+", "_", stem.lower())
    s = re.sub(r"_+", "_", s).strip("_")
    return s or "untitled"


def _is_section_marker(
    question: str,
    answer: str,
    context: str,
    metadata_values: list[str],
) -> bool:
    """Conservative section-marker detector. See the module docstring's
    ARCHITECTURAL DECISION block for the rule and its rationale.

    Returns True only if every condition holds; any one failing means
    "treat this as a normal Q&A row, not a section divider".
    """
    if not question:
        return False
    if answer or context:
        return False
    if any(v for v in metadata_values):
        return False
    if question.rstrip().endswith("?"):
        return False
    word_count = len(question.split())
    return 1 <= word_count <= 4


def _validate_config(config: dict) -> None:
    """Refuse to load if the config is missing required fields.

    A misconfigured config produces wrong data silently — wrong column
    loaded as question, wrong starting row, mis-attributed metadata —
    which is the worst class of bug because it propagates all the way
    into retrieval results before anyone notices. Failing loud at the
    front door is cheaper than debugging a corrupted vector store.
    """
    required = {"source_file", "sheet", "header_row", "columns"}
    missing = required - set(config)
    if missing:
        raise ValueError(
            f"Config is missing required fields: {sorted(missing)}. "
            "Older configs (pre-header_row) need to be regenerated by "
            "re-running profile_excel.py."
        )
    columns = config["columns"]
    for required_role in ("question", "answer"):
        if required_role not in columns:
            raise ValueError(
                f"Config 'columns' must define '{required_role}'. "
                f"Got: {sorted(columns)}. Re-profile this file."
            )


# ─── Public API ─────────────────────────────────────────────────────────
def load_excel(path: str | Path, config: dict) -> list[Row]:
    """Load one Excel RFI file into a list of Row dataclasses.

    Args:
        path: filesystem path to the .xlsx file. The basename must match
              config['source_file'] — a mismatch is a wiring error
              (the wrong config was paired with the file) and raises
              ValueError rather than silently loading the wrong data.
        config: dict loaded from a `config_rfi_<slug>.json`. Required
                keys: source_file, sheet, header_row, columns.
                See profile_excel.build_config() for the full schema.

    Returns:
        A list of Row, one per non-empty data row. The list order
        matches the Excel sheet order. Rows where BOTH the question
        and answer cells are empty are skipped. Each Row's pair_id
        encodes the original Excel row number for traceability.

    Raises:
        ValueError: if the config is malformed, the file/config don't
                    match by basename, or the named sheet doesn't exist.
    """
    path = Path(path)
    _validate_config(config)
    if path.name != config["source_file"]:
        raise ValueError(
            f"path basename '{path.name}' does not match config "
            f"source_file '{config['source_file']}'. Refusing to load "
            "to avoid silent misattribution."
        )

    sheet_name = config["sheet"]
    header_row = int(config["header_row"])
    columns = config["columns"]
    metadata_field_names = list(config.get("metadata_fields", []))
    client = config.get("client")
    date = config.get("date")

    wb = load_workbook(str(path), data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found in {path.name}. "
            f"Available: {wb.sheetnames}."
        )
    ws = wb[sheet_name]

    # Resolve column letters → 1-based indexes once, up front. Cheaper
    # than re-resolving per row and surfaces config errors immediately
    # (e.g. an invalid column letter raises before we open the file).
    q_col = column_index_from_string(columns["question"])
    a_col = column_index_from_string(columns["answer"])
    c_col = (column_index_from_string(columns["context"])
             if "context" in columns else None)
    metadata_cols: dict[str, int] = {
        name: column_index_from_string(columns[name])
        for name in metadata_field_names
        if name in columns
    }

    slug = _slug(path.stem)
    max_row = ws.max_row or 0
    first_data_row = header_row + 1
    if first_data_row > max_row:
        return []

    # current_section is updated whenever we hit a section-marker row;
    # it propagates into the metadata of every subsequent loaded row
    # until the next marker. Reset to None on a fresh load.
    current_section: str | None = None

    rows: list[Row] = []
    for offset, row_values in enumerate(ws.iter_rows(
            min_row=first_data_row,
            max_row=max_row,
            values_only=True)):
        excel_row = first_data_row + offset
        question = _cell_at(row_values, q_col)
        answer = _cell_at(row_values, a_col)
        # No question = no Q&A row, regardless of what the answer
        # cell holds. See the ARCHITECTURAL DECISION block above on
        # why this is stricter than "both empty".
        if not question:
            continue
        context = _cell_at(row_values, c_col) if c_col else ""

        row_metadata_values = [
            _cell_at(row_values, col_idx) for col_idx in metadata_cols.values()
        ]
        if _is_section_marker(question, answer, context, row_metadata_values):
            current_section = question
            continue

        metadata: dict[str, Any] = {}
        for name, col_idx in metadata_cols.items():
            v = _cell_at(row_values, col_idx)
            if v:
                metadata[name] = v
        # Inferred section is attached only if the file has no
        # explicit `section` column (which would already be in
        # metadata by this point). Explicit data always wins.
        if current_section and "section" not in metadata:
            metadata["section"] = current_section
        if client:
            metadata["client"] = client
        if date:
            metadata["date"] = date

        rows.append(Row(
            question=question,
            answer=answer,
            context=context if context else None,
            metadata=metadata,
            source_format="excel",
            source_file=path.name,
            pair_id=f"{slug}_row_{excel_row}",
        ))

    wb.close()
    return rows
